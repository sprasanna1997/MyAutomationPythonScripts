import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
options=webdriver.ChromeOptions()
options.add_experimental_option('detach',True)
driver=webdriver.Chrome(options=options)
driver.get('https://www.tutorialspoint.com/selenium/practice/selenium_automation_practice.php')
name='Prasanna'
print('New Hello World Komalii')
driver.close()
path='D:/Automation/Openpyxl.xlsx'
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name('Sheet1')
rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        res=sheet.cell(row=r,column=c).value
        print(res,end="                     ")

    print()

path='D:/Automation/Test.xlsx'
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name('Sheet1')

for r in range(1,6):
    for c in range(1,5):
        res=sheet.cell(row=r,column=c).value='Automation is Very Good'
        print(res)
workbook.save(path)

logging.basicConfig(filename='log.txt',
                    format='%(asctime)s: %(levelname)s: %(message)s ',
                    datefmt='%d %m %y %I:%M:%S %p',
                    level=logging.DEBUG)

logging.debug("This is a Debug")
logging.info("This is a Info")
logging.warning("This is a Warning")
logging.error("This is a Error")
logging.critical("This is a Critical")
